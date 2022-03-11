import logging
import traceback
from openpyxl import Workbook
import csv
import io
from flask import request, Response, json
from werkzeug.datastructures import FileStorage
from flask_restx import Resource, Namespace
import threading
from config import app_config
from common.utils.response import make_json_response
from dao.dao_base import DAOBaseClass
from dao.dao_function import DAOFunction
from dao import init_data
from service.resources.service_resources import ResourcesService

logger = logging.getLogger(app_config.LOG)

IMPORT = Namespace(name='IMPORT', description='データベーステーブルのImport用API。')

sem = threading.Semaphore()


@IMPORT.route('/rules')
class ImportDefault(Resource):
    parser = IMPORT.parser()
    parser.add_argument('files', type=FileStorage, location='files', required=True, help='FILE名')

    @IMPORT.expect(parser)
    @IMPORT.response(200, 'Success')
    @IMPORT.response(400, 'Bad Request')
    @IMPORT.response(500, 'Internal Server Error')
    def post(self):
        """
        データベース テーブルのImportを担当する。
        Import失敗時本来データで復旧させる。
        """
        logger.info(str(request))

        args = self.parser.parse_args()

        logger.debug(f'args : {args}')
        # f = request.files['file']
        import_file = args['files']

        sem.acquire()
        try:
            if '.xlsx' not in import_file.filename.lower():
                sem.release()
                logger.debug(f'not a excel file : {import_file.filename.lower()}')
                return make_json_response(status=400, msg='Accept only Excel(.xlsx) file format.')

            dao = DAOBaseClass()
            resp_form = dao.export_tables_from_schema(schema_list=app_config.SCHEMA_EXPORT_LIST,
                                                      omit_tb_list=app_config.EXPORT_OMIT_TBL_LIST)
            if not resp_form.res:
                sem.release()
                return make_json_response(status=400, msg=f'Fail export table from schema() : {resp_form.msg}')

            backup_archive = io.BytesIO()
            wb = Workbook()
            sheet_idx = 0
            for tb_name, buffer in resp_form.data.items():
                ws = wb.create_sheet(title=tb_name, index=sheet_idx)
                reader = csv.reader(buffer.getvalue().splitlines(), delimiter='\t')
                for row in reader:
                    ws.append(row)
                sheet_idx += 1

            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            wb.save(backup_archive)

            resp_form = dao.remove_all_records(schema_list=app_config.SCHEMA_EXPORT_LIST,
                                               omit_tb_list=app_config.EXPORT_OMIT_TBL_LIST)
            if not resp_form.res:
                init_data(file=backup_archive)
                sem.release()
                return make_json_response(status=400, msg=f'Fail remove all records(): {resp_form.msg}')

            resp_form = init_data(file=import_file)
            if not resp_form.res:
                dao.remove_all_records(schema_list=app_config.SCHEMA_EXPORT_LIST,
                                       omit_tb_list=app_config.EXPORT_OMIT_TBL_LIST)
                init_data(file=backup_archive)
                sem.release()
                return make_json_response(status=400, msg=f'Fail import data : {resp_form.msg}')

            sem.release()
            return Response(status=200)

        except Exception as e:
            sem.release()
            logger.error(str(e))
            logger.error(traceback.format_exc())
            return make_json_response(status=500, msg=str(e))


@IMPORT.route('/function')
class ImportDefault(Resource):
    parser = IMPORT.parser()
    parser.add_argument('files', type=FileStorage, location='files', required=True, help='FILE名')

    @IMPORT.expect(parser)
    @IMPORT.response(200, 'Success')
    @IMPORT.response(400, 'Bad Request')
    def post(self):
        """
        Import Function Information
        """
        logger.info(str(request))

        args = self.parser.parse_args()

        logger.debug(f'args : {args}')
        import_file = args['files']

        sem.acquire()

        try:
            if '.json' not in import_file.filename.lower():
                sem.release()
                logger.debug(f'not a json file : {import_file.filename.lower()}')
                return make_json_response(status=400, msg='Accept only JSON(.json) file format.')

            func_list = json.load(import_file)

            dao_func = DAOFunction()
            resources = ResourcesService()

            for function in func_list:
                category = function['category']
                func = function['func']
                convert = function['convert']
                analysis = function['analysis']
                visualization = function['visualization']

                resp_form = dao_func.insert_category_info(category)
                if not resp_form.res:
                    sem.release()
                    return make_json_response(status=400, msg=resp_form.msg)

                category_id = resp_form.data

                resp_form = dao_func.insert_func_info({**func,
                                                       'category_id': category_id,
                                                       'analysis_type': analysis['type']})
                if not resp_form.res:
                    sem.release()
                    return make_json_response(status=400, msg=resp_form.msg)

                func_id = resp_form.data

                if func['source_type'] == 'local':
                    resp_form = dao_func.insert_convert_script(convert['script'], func_id=func_id)
                    if not resp_form.res:
                        resources.delete_id(func_id=func_id)
                        sem.release()
                        return make_json_response(status=400, msg=resp_form.msg)

                resp_form = dao_func.insert_analysis_info(analysis, func_id)
                if not resp_form.res:
                    resources.delete_id(func_id=func_id)
                    sem.release()
                    return make_json_response(status=400, msg=resp_form.msg)

                resp_form = dao_func.insert_visual_info(visualization, func_id)
                if not resp_form.res:
                    resources.delete_id(func_id=func_id)
                    sem.release()
                    return make_json_response(status=400, msg=resp_form.msg)

            sem.release()
            return Response(status=200)

        except Exception as e:
            sem.release()
            logger.error(str(e))
            logger.error(traceback.format_exc())
            return make_json_response(status=400, msg=str(e))
