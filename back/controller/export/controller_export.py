import base64
import logging
import traceback
import io
import os
import zipfile
import csv
from openpyxl import Workbook
from flask import request, Response, json
from flask_restx import Resource, Namespace, fields
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename

from config import app_config
from common.utils.response import make_json_response, ResponseForm
from dao.dao_base import DAOBaseClass
from dao.dao_function import DAOFunction
from service.overlay.service_overlay_correction import ServiceCorrection
from service.overlay.service_overlay_AdcMeasurement import ServiceAdcMeasurement

logger = logging.getLogger(app_config.LOG)

EXPORT = Namespace(name='EXPORT', description='ログ変換に関連した設定情報のExport用API')

export_response = EXPORT.model('export_table_response', {
    'version': fields.String(description='description', example='example')
})


@EXPORT.route('/rules')
class ExportTableDefault(Resource):
    @EXPORT.response(200, 'Success')
    @EXPORT.response(500, 'Internal Server Error')
    def get(self):
        """
        CNVBASE SchemaのテーブルデータをZIP圧縮ファイルで生成する。
        """
        logger.info(str(request))

        try:
            dao = DAOBaseClass()

            resp_form = dao.export_tables_from_schema(schema_list=app_config.SCHEMA_EXPORT_LIST,
                                                      omit_tb_list=app_config.EXPORT_OMIT_TBL_LIST)
            if resp_form.res:
                try:
                    archive = io.BytesIO()

                    wb = Workbook()
                    sheet_idx = 0
                    for tb_name, buffer in resp_form.data.items():
                        ws = wb.create_sheet(title=tb_name, index=sheet_idx)
                        reader = csv.reader(buffer.getvalue().splitlines(), delimiter='\t')
                        for row in reader:
                            ws.append(row)
                        sheet_idx += 1

                    if 'Sheet' in wb.sheetnames:
                        wb.remove(wb['Sheet'])
                    wb.save(archive)

                    # with zipfile.ZipFile(archive, 'w', zipfile.ZIP_DEFLATED) as zip_archive:
                    #     for key, buffer in resp_form.data.items():
                    #         zip_archive.writestr(key+'.csv', buffer.getvalue())
                except Exception as e:
                    logger.error(str(e))
                    return Response(status=500)

                return Response(archive.getvalue(), headers={
                    # 'Content-Type': 'application/zip',
                    # 'Content-Disposition': 'attachment; filename=%s;' % 'export.zip'
                    'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    'Content-Disposition': f'attachment; filename=export.xlsx'
                })
            else:
                return make_json_response(status=500, msg=resp_form.msg)
        except Exception as e:
            logger.error(str(e))
            logger.error(traceback.format_exc())
            return make_json_response(status=500, msg=str(e))


@EXPORT.route('')
class ExportDefault(Resource):
    parser = EXPORT.parser()
    parser.add_argument('files', type=FileStorage, location='files', required=True, help='FILE名', action='append')

    @EXPORT.expect(parser)
    @EXPORT.doc(model=export_response)
    @EXPORT.response(200, 'Success')
    @EXPORT.response(500, 'Internal Server Error')
    def post(self):
        """
        tableデータとグラフ イメージのExportを実施する。
        """
        logger.info(str(request))

        # files = request.form.getlist('files')
        # print(files)
        args = self.parser.parse_args()
        logger.debug(f'args : {args}')
        files = args['files']

        try:
            archive = io.BytesIO()

            with zipfile.ZipFile(archive, 'w', zipfile.ZIP_DEFLATED) as zip_archive:
                for file in files:
                    filename = file.filename
                    raw = file.read()

                    filename = filename.split(sep='.')
                    folder = filename[0]
                    extension = filename[-1]
                    filename = '.'.join(filename[1:]).replace('/', '_')

                    if extension == 'png':
                        raw = raw.replace(b'data:image/png;base64,', b'')
                        raw = base64.b64decode(raw)

                    zip_archive.writestr(os.path.join(folder, filename), raw)

            # with open('my_file.zip', 'wb') as f_out:
            #     f_out.write(archive.getvalue())

        except Exception as e:
            logger.error(str(e))
            logger.error(traceback.format_exc())
            return Response(status=500)

        return Response(archive.getvalue(), headers={
            'Content-Type': 'application/zip',
            'Content-Disposition': 'attachment; filename=%s;' % 'export.zip'
        })


@EXPORT.route('/function')
class ExportTableDefault(Resource):
    parser = EXPORT.parser()
    parser.add_argument('func_id', required=True, help='Function ID. Multiple Id allowed.')

    @EXPORT.expect(parser)
    @EXPORT.response(200, 'Success')
    @EXPORT.response(400, 'Internal Server Error')
    def get(self):
        """
        Export Function Information as Json Format
        """
        logger.info(str(request))

        try:
            param_dict = request.args.to_dict(flat=False)
            func_id_list = param_dict['func_id']

            resp_form = self.get_func_export_info(func_id_list)
            if not resp_form.res:
                return make_json_response(status=400, msg=resp_form.msg)

            data_json = json.dumps(resp_form.data, indent=4, sort_keys=True, ensure_ascii=False)

            return Response(data_json, headers={
                'Content-Type': 'application/json; charset=utf-8',
                # 'Content-Disposition': 'attachment; filename=%s;' % 'export.zip'
                'Content-Disposition': f'attachment; filename=function_export.json'
            })
        except Exception as e:
            logger.error(str(e))
            logger.error(traceback.format_exc())
            return make_json_response(status=400, msg=str(e))

    def get_func_export_info(self, id_list):
        if len(id_list) == 0:
            return ResponseForm(res=False, msg='Function ID Empty.')

        try:
            ret_list = []
            for func_id in id_list:
                ret_dict = dict()
                dao_func = DAOFunction()

                resp_form = dao_func.get_category_info(func_id)
                if not resp_form.res:
                    logger.info(f'get_category_info() : {resp_form.msg}')
                    continue

                ret_dict['category'] = resp_form.data

                resp_form = dao_func.get_function_info(func_id)
                if not resp_form.res:
                    logger.info(f'get_function_info() : {resp_form.msg}')
                    continue

                analysis_type = resp_form.data.pop('analysis_type')
                ret_dict['func'] = resp_form.data

                if analysis_type == 'multi':
                    logger.info(f'Not support exporting multi function.')
                    continue

                resp_form = dao_func.get_source_info(func_id)
                if not resp_form.res:
                    logger.info(f'get_source_info() : {resp_form.msg}')
                    continue

                ret_dict['func']['info'] = resp_form.data

                resp_form = dao_func.get_script(table='analysis.preprocess_script', func_id=func_id)
                if not resp_form.res:
                    preprocess_script = {'file_name': None, 'use_script': False}
                else:
                    preprocess_script = resp_form.data

                ret_dict['func']['script'] = preprocess_script

                ret_dict['convert'] = dict()
                if ret_dict['func']['source_type'] == 'local':
                    resp_form = dao_func.get_script(table='analysis.convert_script', func_id=func_id)
                    if not resp_form.res:
                        convert_script = {'file_name': None, 'use_script': False}
                    else:
                        convert_script = resp_form.data

                    ret_dict['convert']['script'] = convert_script

                analysis_items = list()
                resp_form = dao_func.get_analysis_items_info(func_id)
                if resp_form.res:
                    analysis_items = resp_form.data

                ret_dict['analysis'] = dict()
                ret_dict['analysis']['type'] = analysis_type
                ret_dict['analysis']['setting'] = dict()
                ret_dict['analysis']['setting']['items'] = analysis_items

                resp_form = dao_func.get_filter_default_info(func_id)
                if not resp_form.res:
                    filter_default = []
                else:
                    filter_default = resp_form.data

                ret_dict['analysis']['setting']['filter_default'] = filter_default

                resp_form = dao_func.get_aggregation_default_info(func_id)
                if not resp_form.res:
                    aggregation_default = dict()
                else:
                    aggregation_default = resp_form.data

                ret_dict['analysis']['setting']['aggregation_default'] = aggregation_default

                resp_form = dao_func.get_script(table='analysis.analysis_script', func_id=func_id)
                if not resp_form.res:
                    analysis_script = {'db_id': None, 'sql': None, 'file_name': None, 'use_script': False}
                else:
                    analysis_script = resp_form.data

                ret_dict['analysis']['script'] = analysis_script

                resp_form = dao_func.get_function_graph_type(func_id)
                if not resp_form.res:
                    continue

                ret_dict['visualization'] = dict()
                ret_dict['visualization']['function_graph_type'] = resp_form.data

                resp_form = dao_func.get_visualization_default_info(func_id)
                if not resp_form.res:
                    items = []
                else:
                    items = resp_form.data

                ret_dict['visualization']['items'] = items

                ret_list.append(ret_dict)

            if len(ret_list) == 0:
                return ResponseForm(res=False, msg='Empty.')
            else:
                return ResponseForm(res=True, data=ret_list)

        except Exception as e:
            logger.error(str(e))
            logger.error(traceback.format_exc())
            return ResponseForm(res=False, msg=str(e))


@EXPORT.route('/overlay/<string:category>/<string:rid>')
@EXPORT.param('category', 'Category')
@EXPORT.param('rid', 'Request ID')
class ExportOverlay(Resource):
    parser = EXPORT.parser()
    parser.add_argument('files', type=FileStorage, location='files', required=True, help='graph files', action='append')
    parser.add_argument('setting', type=FileStorage, location='files', required=True, help='setting information')

    @EXPORT.expect(parser)
    @EXPORT.response(200, 'Success')
    @EXPORT.response(400, 'Bad Request')
    def post(self, category, rid):
        """
        tableデータとグラフ イメージのExportを実施する。
        """
        logger.info(str(request))

        args = self.parser.parse_args()
        logger.debug(f'args : {args}')
        files = args['files']
        setting_json = args['setting']

        try:
            adc_meas_df = None
            correction_df = None

            setting_dict = json.load(setting_json)

            service_adc_meas = ServiceAdcMeasurement()
            param = dict()
            param['period'] = setting_dict['period']
            param['job'] = setting_dict['job']
            param['lot_id'] = setting_dict['lot_id']
            param['rid'] = rid
            param['cp_vs'] = setting_dict.pop('cp_vs')

            adc_meas_df, cp_vs_included = service_adc_meas.get_adc_meas_data(param)

            if category == 'correction':
                service_correction = ServiceCorrection()
                correction_df = service_correction.get_correction_data(param, to_df=True)

            archive = io.BytesIO()

            with zipfile.ZipFile(archive, 'w', zipfile.ZIP_DEFLATED) as zip_archive:
                for file in files:
                    filename = file.filename
                    raw = file.read()

                    filename_sep = filename.split(sep='.')
                    if len(filename_sep) > 2:
                        folder = filename_sep[0]
                        extension = filename_sep[-1]
                        filename = '.'.join(filename_sep[1:]).replace('/', '_')
                    else:
                        folder = None
                        extension = filename_sep[-1]

                    if extension == 'png':
                        raw = raw.replace(b'data:image/png;base64,', b'')
                        raw = base64.b64decode(raw)

                    if folder is None:
                        zip_archive.writestr(filename, raw)
                    else:
                        zip_archive.writestr(os.path.join(folder, filename), raw)

                if adc_meas_df is not None:
                    zip_archive.writestr('AdcMeasurement.csv', adc_meas_df.to_csv(header=True, index=False))

                if correction_df is not None:
                    zip_archive.writestr('CorrectionComponentsMap.csv', correction_df.to_csv(header=True, index=False))

                zip_archive.writestr('setting.json', json.dumps(setting_dict, indent=4, sort_keys=True, ensure_ascii=False))

        except Exception as e:
            logger.error(str(e))
            logger.error(traceback.format_exc())
            return make_json_response(status=500, msg=str(e))

        return Response(archive.getvalue(), headers={
            'Content-Type': 'application/zip',
            'Content-Disposition': 'attachment; filename=%s;' % 'export.zip'
        })
